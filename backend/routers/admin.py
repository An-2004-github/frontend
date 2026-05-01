from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File
import io
from sqlalchemy import text
from database import engine
from auth import get_current_user
from pydantic import BaseModel
from typing import Optional
from email_service import send_booking_confirmation_email
from rank_utils import get_rank, get_cashback_rate
from routers.notifications import create_notification

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Kiểm tra quyền admin ────────────────────────────────────────
def get_admin_user(user_id: int = Depends(get_current_user)):
    with engine.connect() as conn:
        user = conn.execute(
            text("SELECT role FROM users WHERE user_id = :uid"),
            {"uid": user_id}
        ).fetchone()
        if not user or dict(user._mapping).get("role") != "ADMIN":
            raise HTTPException(403, "Không có quyền truy cập")
    return user_id


# ── STATS ───────────────────────────────────────────────────────
@router.get("/stats")
def get_stats(admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        total_users    = conn.execute(text("SELECT COUNT(*) FROM users")).scalar()
        total_bookings = conn.execute(text("SELECT COUNT(*) FROM bookings")).scalar()
        confirmed      = conn.execute(text("SELECT COUNT(*) FROM bookings WHERE status='confirmed'")).scalar()
        pending        = conn.execute(text("SELECT COUNT(*) FROM bookings WHERE status='pending'")).scalar()
        revenue        = conn.execute(text("SELECT COALESCE(SUM(final_amount),0) FROM bookings WHERE status='confirmed'")).scalar()
        total_hotels   = conn.execute(text("SELECT COUNT(*) FROM hotels")).scalar()
        total_flights  = conn.execute(text("SELECT COUNT(*) FROM flights")).scalar()
        total_buses    = conn.execute(text("SELECT COUNT(*) FROM buses")).scalar()
        total_trains   = conn.execute(text("SELECT COUNT(*) FROM trains")).scalar()

        # Doanh thu 7 ngày gần nhất
        revenue_7d = conn.execute(text("""
            SELECT DATE(booking_date) as day, COALESCE(SUM(final_amount),0) as total
            FROM bookings
            WHERE status='confirmed' AND booking_date >= DATE_SUB(NOW(), INTERVAL 7 DAY)
            GROUP BY DATE(booking_date)
            ORDER BY day
        """)).fetchall()

    return {
        "total_users": total_users,
        "total_bookings": total_bookings,
        "confirmed_bookings": confirmed,
        "pending_bookings": pending,
        "total_revenue": float(revenue),
        "total_hotels": total_hotels,
        "total_flights": total_flights,
        "total_buses": total_buses,
        "total_trains": total_trains,
        "revenue_7d": [{"day": str(r.day), "total": float(r.total)} for r in revenue_7d],
    }


# ── REVENUE CHART ───────────────────────────────────────────────
@router.get("/revenue")
def get_revenue(period: str = "7d", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        if period == "monthly":
            rows = conn.execute(text("""
                SELECT DATE_FORMAT(booking_date, '%Y-%m') as label,
                       COALESCE(SUM(final_amount), 0) as total
                FROM bookings
                WHERE status = 'confirmed'
                  AND booking_date >= DATE_SUB(NOW(), INTERVAL 12 MONTH)
                GROUP BY DATE_FORMAT(booking_date, '%Y-%m')
                ORDER BY label
            """)).fetchall()
        else:
            rows = conn.execute(text("""
                SELECT DATE(booking_date) as label,
                       COALESCE(SUM(final_amount), 0) as total
                FROM bookings
                WHERE status = 'confirmed'
                  AND booking_date >= DATE_SUB(NOW(), INTERVAL 7 DAY)
                GROUP BY DATE(booking_date)
                ORDER BY label
            """)).fetchall()
    return [{"label": str(r.label), "total": float(r.total)} for r in rows]


# ── USERS ────────────────────────────────────────────────────────
@router.get("/users")
def get_all_users(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE full_name LIKE :s OR email LIKE :s OR phone LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(
            f"SELECT user_id, full_name, email, phone, role, wallet, provider, created_at FROM users {where} ORDER BY created_at DESC, user_id DESC LIMIT :limit OFFSET :skip"
        ), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.put("/users/{user_id}/role")
def update_user_role(user_id: int, role: str, admin_id: int = Depends(get_admin_user)):
    if role not in ("USER", "ADMIN"):
        raise HTTPException(400, "Role không hợp lệ")
    with engine.begin() as conn:
        conn.execute(text("UPDATE users SET role = :role WHERE user_id = :uid"), {"role": role, "uid": user_id})
    return {"message": "Cập nhật role thành công"}


@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin_id: int = Depends(get_admin_user)):
    if user_id == admin_id:
        raise HTTPException(400, "Không thể xóa chính mình")
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM users WHERE user_id = :uid"), {"uid": user_id})
    return {"message": "Xóa người dùng thành công"}


class UserCreateRequest(BaseModel):
    full_name: str
    email: str
    password: str
    phone: Optional[str] = None
    role: Optional[str] = "USER"


class UserUpdateRequest(BaseModel):
    full_name: str
    email: str
    phone: Optional[str] = None
    role: Optional[str] = "USER"
    new_password: Optional[str] = None  # nếu để trống thì không đổi mật khẩu


@router.put("/users/{user_id}")
def admin_update_user(user_id: int, data: UserUpdateRequest, admin_id: int = Depends(get_admin_user)):
    from auth import hash_password
    if data.role not in ("USER", "ADMIN"):
        raise HTTPException(400, "Role không hợp lệ")
    with engine.begin() as conn:
        # Kiểm tra email trùng với user khác
        existing = conn.execute(
            text("SELECT user_id FROM users WHERE email = :email AND user_id != :uid"),
            {"email": data.email, "uid": user_id}
        ).fetchone()
        if existing:
            raise HTTPException(400, "Email đã được dùng bởi tài khoản khác")

        fields = {
            "full_name": data.full_name,
            "email": data.email,
            "phone": data.phone,
            "role": data.role,
            "uid": user_id,
        }
        set_clause = "full_name=:full_name, email=:email, phone=:phone, role=:role"

        if data.new_password:
            if len(data.new_password) < 6:
                raise HTTPException(400, "Mật khẩu phải có ít nhất 6 ký tự")
            fields["password_hash"] = hash_password(data.new_password)
            set_clause += ", password_hash=:password_hash"

        conn.execute(
            text(f"UPDATE users SET {set_clause} WHERE user_id=:uid"),
            fields
        )
    return {"message": "Cập nhật người dùng thành công"}


@router.post("/users")
def admin_create_user(data: UserCreateRequest, admin_id: int = Depends(get_admin_user)):
    from auth import hash_password
    if data.role not in ("USER", "ADMIN"):
        raise HTTPException(400, "Role không hợp lệ")
    with engine.begin() as conn:
        existing = conn.execute(
            text("SELECT user_id FROM users WHERE email = :email"), {"email": data.email}
        ).fetchone()
        if existing:
            raise HTTPException(400, "Email đã tồn tại")
        hashed = hash_password(data.password)
        result = conn.execute(text("""
            INSERT INTO users (full_name, email, password_hash, phone, role, wallet, provider)
            VALUES (:name, :email, :password, :phone, :role, 0, 'local')
        """), {
            "name": data.full_name, "email": data.email, "password": hashed,
            "phone": data.phone, "role": data.role,
        })
    return {"user_id": result.lastrowid, "message": "Tạo người dùng thành công"}


# ── BOOKINGS ─────────────────────────────────────────────────────
@router.get("/bookings")
def get_all_bookings(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        having = ""
        if search:
            having = "HAVING user_name LIKE :s OR user_email LIKE :s OR entity_name LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT
                b.booking_id, b.booking_date, b.status, b.total_price, b.final_amount,
                u.full_name AS user_name, u.email AS user_email,
                MAX(bi.entity_type) AS entity_type,
                MAX(CASE bi.entity_type
                    WHEN 'room'   THEN CONCAT(h.name, ' - ', rt.name)
                    WHEN 'flight' THEN CONCAT(f.airline, ': ', f.from_city, ' \u2192 ', f.to_city)
                    WHEN 'bus'    THEN CONCAT(bs.company, ': ', bs.from_city, ' \u2192 ', bs.to_city)
                    WHEN 'train'  THEN CONCAT('Tàu ', t.train_code, ': ', t.from_city, ' \u2192 ', t.to_city)
                    ELSE CONCAT(bi.entity_type, ' #', bi.entity_id)
                END) AS entity_name
            FROM bookings b
            JOIN users u ON u.user_id = b.user_id
            LEFT JOIN booking_items bi ON bi.booking_id = b.booking_id
            LEFT JOIN room_types rt ON rt.room_type_id = bi.entity_id AND bi.entity_type = 'room'
            LEFT JOIN hotels h ON h.hotel_id = rt.hotel_id
            LEFT JOIN flights f ON f.flight_id = bi.entity_id AND bi.entity_type = 'flight'
            LEFT JOIN buses bs ON bs.bus_id = bi.entity_id AND bi.entity_type = 'bus'
            LEFT JOIN trains t ON t.train_id = bi.entity_id AND bi.entity_type = 'train'
            GROUP BY b.booking_id, b.booking_date, b.status, b.total_price, b.final_amount,
                     u.full_name, u.email
            {having}
            ORDER BY b.booking_date DESC, b.booking_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/bookings/{booking_id}/detail")
def get_booking_detail_admin(booking_id: int, admin_id: int = Depends(get_admin_user)):
    import json as _json
    with engine.connect() as conn:
        booking = conn.execute(
            text("SELECT b.*, u.full_name AS user_name, u.email AS user_email FROM bookings b JOIN users u ON u.user_id = b.user_id WHERE b.booking_id = :id"),
            {"id": booking_id}
        ).fetchone()
        if not booking:
            raise HTTPException(404, "Booking không tồn tại")
        result = dict(booking._mapping)

        # Items
        items = conn.execute(text("""
            SELECT bi.*,
                CASE bi.entity_type
                    WHEN 'room'   THEN CONCAT(h.name, ' - ', rt.name)
                    WHEN 'flight' THEN CONCAT(f.airline, ': ', f.from_city, ' → ', f.to_city)
                    WHEN 'bus'    THEN CONCAT(bs.company, ': ', bs.from_city, ' → ', bs.to_city)
                    WHEN 'train'  THEN CONCAT('Tàu ', t.train_code, ': ', t.from_city, ' → ', t.to_city)
                    ELSE CONCAT(bi.entity_type, ' #', bi.entity_id)
                END AS entity_name
            FROM booking_items bi
            LEFT JOIN room_types rt ON rt.room_type_id = bi.entity_id AND bi.entity_type = 'room'
            LEFT JOIN hotels h ON h.hotel_id = rt.hotel_id
            LEFT JOIN flights f ON f.flight_id = bi.entity_id AND bi.entity_type = 'flight'
            LEFT JOIN buses bs ON bs.bus_id = bi.entity_id AND bi.entity_type = 'bus'
            LEFT JOIN trains t ON t.train_id = bi.entity_id AND bi.entity_type = 'train'
            WHERE bi.booking_id = :id
        """), {"id": booking_id}).fetchall()
        result["items"] = [dict(i._mapping) for i in items]

        # Payment transactions
        payments = conn.execute(text("""
            SELECT method, amount, status, paid_at, transaction_ref
            FROM payment_transactions WHERE booking_id = :id ORDER BY paid_at ASC
        """), {"id": booking_id}).fetchall()
        result["payments"] = [dict(p._mapping) for p in payments]

        # Modification history (full)
        mods = conn.execute(text("""
            SELECT m.*, u2.full_name AS acted_by
            FROM booking_modifications m
            LEFT JOIN users u2 ON u2.user_id = m.user_id
            WHERE m.booking_id = :id ORDER BY m.created_at ASC
        """), {"id": booking_id}).fetchall()
        result["modifications"] = [dict(m._mapping) for m in mods]

        return result


@router.put("/bookings/{booking_id}/status")
def update_booking_status(booking_id: int, status: str, background_tasks: BackgroundTasks, admin_id: int = Depends(get_admin_user)):
    if status not in ("pending", "confirmed", "cancelled"):
        raise HTTPException(400, "Status không hợp lệ")
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE bookings SET status = :status WHERE booking_id = :id"),
            {"status": status, "id": booking_id}
        )

        # Gửi email xác nhận khi admin duyệt thanh toán QR
        if status == "confirmed":
            row = conn.execute(text("""
                SELECT
                    u.email, u.full_name,
                    b.final_amount, b.booking_date,
                    bi.entity_type, bi.check_in_date, bi.check_out_date,
                    bi.quantity, bi.guests,
                    CASE bi.entity_type
                        WHEN 'room'   THEN CONCAT(h.name, ' - ', rt.name)
                        WHEN 'flight' THEN CONCAT(f.airline, ': ', f.from_city, ' → ', f.to_city)
                        WHEN 'bus'    THEN CONCAT(bs.company, ': ', bs.from_city, ' → ', bs.to_city)
                        ELSE CONCAT(bi.entity_type, ' #', bi.entity_id)
                    END AS service_name
                FROM bookings b
                JOIN users u ON u.user_id = b.user_id
                JOIN booking_items bi ON bi.booking_id = b.booking_id
                LEFT JOIN room_types rt ON rt.room_type_id = bi.entity_id AND bi.entity_type = 'room'
                LEFT JOIN hotels h ON h.hotel_id = rt.hotel_id
                LEFT JOIN flights f ON f.flight_id = bi.entity_id AND bi.entity_type = 'flight'
                LEFT JOIN buses bs ON bs.bus_id = bi.entity_id AND bi.entity_type = 'bus'
                WHERE b.booking_id = :bid LIMIT 1
            """), {"bid": booking_id}).fetchone()

            if row:
                info = dict(row._mapping)
                background_tasks.add_task(
                    send_booking_confirmation_email,
                    email=info["email"],
                    name=info["full_name"] or "Quý khách",
                    booking_id=booking_id,
                    service_name=info["service_name"] or f"Dịch vụ #{booking_id}",
                    amount=float(info["final_amount"]),
                    entity_type=info["entity_type"] or "",
                    check_in=str(info["check_in_date"]) if info["check_in_date"] else None,
                    check_out=str(info["check_out_date"]) if info["check_out_date"] else None,
                    booking_date=str(info["booking_date"])[:10] if info["booking_date"] else None,
                    rooms=int(info["quantity"]) if info.get("quantity") else None,
                    guests=int(info["guests"]) if info.get("guests") else None,
                )

                # Cập nhật rank + cộng cashback vào ví
                booking_user = conn.execute(
                    text("SELECT user_id FROM bookings WHERE booking_id = :bid"),
                    {"bid": booking_id}
                ).fetchone()
                if booking_user:
                    uid = booking_user.user_id
                    total_spent = conn.execute(
                        text("SELECT COALESCE(SUM(final_amount), 0) FROM bookings WHERE user_id = :uid AND status = 'confirmed'"),
                        {"uid": uid}
                    ).scalar()
                    new_rank = get_rank(float(total_spent))
                    conn.execute(
                        text("UPDATE users SET user_rank = :rank WHERE user_id = :uid"),
                        {"rank": new_rank, "uid": uid}
                    )
                    cashback = round(float(info["final_amount"]) * get_cashback_rate(new_rank))
                    if cashback > 0:
                        conn.execute(
                            text("UPDATE users SET wallet = wallet + :amt WHERE user_id = :uid"),
                            {"amt": cashback, "uid": uid}
                        )
                        conn.execute(text("""
                            INSERT INTO wallet_transactions (user_id, amount, type, description, status)
                            VALUES (:uid, :amt, 'cashback', :desc, 'success')
                        """), {"uid": uid, "amt": cashback,
                               "desc": f"Cashback {int(get_cashback_rate(new_rank)*100*10)/10}% đơn #{booking_id}"})

    return {"message": "Cập nhật trạng thái thành công"}


# ── Helper: lưu nhiều ảnh vào bảng images (tối đa 3) ────────────
def _save_image(conn, entity_type: str, entity_id: int, image_url: Optional[str]):
    """Nhận chuỗi URL phân cách bằng dấu phẩy, xóa ảnh cũ rồi insert tối đa 3 ảnh mới."""
    conn.execute(text(
        "DELETE FROM images WHERE entity_type=:et AND entity_id=:eid"
    ), {"et": entity_type, "eid": entity_id})
    if image_url:
        urls = [u.strip() for u in image_url.split(",") if u.strip()][:3]
        for url in urls:
            conn.execute(text(
                "INSERT INTO images (entity_type, entity_id, image_url) VALUES (:et, :eid, :url)"
            ), {"et": entity_type, "eid": entity_id, "url": url})


# ── HOTEL ROOM TYPES ─────────────────────────────────────────────
class RoomTypeRequest(BaseModel):
    name: str
    price_per_night: float
    max_guests: Optional[int] = 2
    image_url: Optional[str] = None


@router.get("/hotels/{hotel_id}/rooms")
def admin_get_rooms(hotel_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT rt.*,
                   GREATEST(0, rt.total_rooms - COALESCE((
                       SELECT COUNT(*)
                       FROM booking_items bi
                       JOIN bookings b ON b.booking_id = bi.booking_id
                       WHERE bi.entity_type = 'room'
                         AND bi.entity_id = rt.room_type_id
                         AND b.status IN ('pending','confirmed')
                         AND bi.check_out_date >= CURDATE()
                   ), 0)) AS available_rooms,
                   (SELECT GROUP_CONCAT(img.image_url ORDER BY img.image_id SEPARATOR ',')
                    FROM images img
                    WHERE img.entity_type='room_type' AND img.entity_id=rt.room_type_id) AS image_url
            FROM room_types rt
            WHERE rt.hotel_id = :hid
            ORDER BY rt.price_per_night ASC
        """), {"hid": hotel_id}).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/hotels/{hotel_id}/rooms")
def admin_create_room(hotel_id: int, data: RoomTypeRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO room_types (hotel_id, name, price_per_night, max_guests)
            VALUES (:hid, :name, :price, :guests)
        """), {
            "hid": hotel_id, "name": data.name, "price": data.price_per_night,
            "guests": data.max_guests,
        })
        room_id = result.lastrowid
        _save_image(conn, "room_type", room_id, data.image_url)
    return {"room_type_id": room_id, "message": "Tạo loại phòng thành công"}


@router.put("/hotels/{hotel_id}/rooms/{room_type_id}")
def admin_update_room(hotel_id: int, room_type_id: int, data: RoomTypeRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE room_types SET name=:name, price_per_night=:price,
            max_guests=:guests
            WHERE room_type_id=:rid AND hotel_id=:hid
        """), {
            "name": data.name, "price": data.price_per_night,
            "guests": data.max_guests, "rid": room_type_id, "hid": hotel_id,
        })
        _save_image(conn, "room_type", room_type_id, data.image_url)
    return {"message": "Cập nhật loại phòng thành công"}


@router.delete("/hotels/{hotel_id}/rooms/{room_type_id}")
def admin_delete_room(hotel_id: int, room_type_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM images WHERE entity_type='room_type' AND entity_id=:id"), {"id": room_type_id})
        conn.execute(text("DELETE FROM room_types WHERE room_type_id=:rid AND hotel_id=:hid"),
                     {"rid": room_type_id, "hid": hotel_id})
    return {"message": "Xóa loại phòng thành công"}


# ── HOTELS ───────────────────────────────────────────────────────
class HotelRequest(BaseModel):
    name: str
    address: str
    destination_id: Optional[int] = None
    description: Optional[str] = None
    amenities: Optional[str] = None
    image_url: Optional[str] = None


@router.get("/hotels")
def admin_get_hotels(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE h.name LIKE :s OR d.city LIKE :s OR h.address LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT h.*, d.city AS dest_city,
                   (SELECT GROUP_CONCAT(img.image_url ORDER BY img.image_id SEPARATOR ',')
                    FROM images img
                    WHERE img.entity_type='hotel' AND img.entity_id=h.hotel_id) AS image_url,
                   (SELECT COALESCE(SUM(rt.total_rooms), 0)
                    FROM room_types rt WHERE rt.hotel_id = h.hotel_id) AS total_rooms,
                   (SELECT GREATEST(0,
                        COALESCE(SUM(rt.total_rooms), 0) - (
                            SELECT COUNT(*)
                            FROM booking_items bi2
                            JOIN bookings b2 ON b2.booking_id = bi2.booking_id
                            JOIN room_types rt2 ON rt2.room_type_id = bi2.entity_id
                            WHERE bi2.entity_type = 'room'
                            AND rt2.hotel_id = h.hotel_id
                            AND b2.status IN ('pending','confirmed')
                            AND bi2.check_out_date >= CURDATE()
                        ))
                    FROM room_types rt WHERE rt.hotel_id = h.hotel_id) AS available_rooms
            FROM hotels h
            LEFT JOIN destinations d ON d.destination_id = h.destination_id
            {where}
            ORDER BY h.hotel_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/hotels")
def admin_create_hotel(data: HotelRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO hotels (name, address, destination_id, description, amenities)
            VALUES (:name, :address, :dest_id, :desc, :amenities)
        """), {
            "name": data.name, "address": data.address, "dest_id": data.destination_id,
            "desc": data.description, "amenities": data.amenities,
        })
        hotel_id = result.lastrowid
        _save_image(conn, "hotel", hotel_id, data.image_url)
    return {"hotel_id": hotel_id, "message": "Tạo khách sạn thành công"}


@router.put("/hotels/{hotel_id}")
def admin_update_hotel(hotel_id: int, data: HotelRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE hotels SET name=:name, address=:address, destination_id=:dest_id,
            description=:desc, amenities=:amenities WHERE hotel_id=:id
        """), {
            "name": data.name, "address": data.address, "dest_id": data.destination_id,
            "desc": data.description, "amenities": data.amenities, "id": hotel_id
        })
        _save_image(conn, "hotel", hotel_id, data.image_url)
    return {"message": "Cập nhật khách sạn thành công"}


@router.delete("/hotels/{hotel_id}")
def admin_delete_hotel(hotel_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM images WHERE entity_type='hotel' AND entity_id=:id"), {"id": hotel_id})
        conn.execute(text("DELETE FROM hotels WHERE hotel_id = :id"), {"id": hotel_id})
    return {"message": "Xóa khách sạn thành công"}


# ── FLIGHTS ──────────────────────────────────────────────────────
class FlightRequest(BaseModel):
    airline: str
    from_city: str
    to_city: str
    depart_time: str
    arrive_time: str
    price: float
    image_url: Optional[str] = None
    status: Optional[str] = "active"


@router.get("/flights")
def admin_get_flights(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE flights SET status = 'completed'
            WHERE status = 'active' AND depart_time < NOW()
        """))
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE f.airline LIKE :s OR f.from_city LIKE :s OR f.to_city LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT f.*,
                   (SELECT img.image_url FROM images img
                    WHERE img.entity_type='flight' AND img.entity_id=f.flight_id LIMIT 1) AS image_url,
                   (SELECT COUNT(*) FROM flight_seats fs
                    WHERE fs.flight_id = f.flight_id AND fs.is_booked = 0) AS avail_total,
                   (SELECT COUNT(*) FROM flight_seats fs
                    WHERE fs.flight_id = f.flight_id) AS total_seats,
                   (SELECT COUNT(*) FROM flight_seats fs
                    WHERE fs.flight_id = f.flight_id AND fs.is_booked = 0 AND fs.seat_class = 'economy') AS avail_economy,
                   (SELECT COUNT(*) FROM flight_seats fs
                    WHERE fs.flight_id = f.flight_id AND fs.is_booked = 0 AND fs.seat_class = 'business') AS avail_business,
                   (SELECT COUNT(*) FROM flight_seats fs
                    WHERE fs.flight_id = f.flight_id AND fs.is_booked = 0 AND fs.seat_class = 'first') AS avail_first
            FROM flights f
            {where}
            ORDER BY f.flight_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/flights")
def admin_create_flight(data: FlightRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO flights (airline, from_city, to_city, depart_time, arrive_time, price)
            VALUES (:airline, :from_city, :to_city, :depart, :arrive, :price)
        """), {
            "airline": data.airline, "from_city": data.from_city, "to_city": data.to_city,
            "depart": data.depart_time, "arrive": data.arrive_time,
            "price": data.price,
        })
        flight_id = result.lastrowid
        _save_image(conn, "flight", flight_id, data.image_url)
    return {"flight_id": flight_id, "message": "Tạo chuyến bay thành công"}


@router.put("/flights/{flight_id}")
def admin_update_flight(flight_id: int, data: FlightRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE flights SET airline=:airline, from_city=:from_city, to_city=:to_city,
            depart_time=:depart, arrive_time=:arrive, price=:price,
            status=:status WHERE flight_id=:id
        """), {
            "airline": data.airline, "from_city": data.from_city, "to_city": data.to_city,
            "depart": data.depart_time, "arrive": data.arrive_time,
            "price": data.price,
            "status": data.status or "active", "id": flight_id
        })
        _save_image(conn, "flight", flight_id, data.image_url)
    return {"message": "Cập nhật chuyến bay thành công"}


@router.delete("/flights/{flight_id}")
def admin_delete_flight(flight_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM images WHERE entity_type='flight' AND entity_id=:id"), {"id": flight_id})
        conn.execute(text("DELETE FROM flight_seats WHERE flight_id = :id"), {"id": flight_id})
        conn.execute(text("DELETE FROM flights WHERE flight_id = :id"), {"id": flight_id})
    return {"message": "Xóa chuyến bay thành công"}


# ── BUSES ────────────────────────────────────────────────────────
class BusRequest(BaseModel):
    company: str
    from_city: str
    to_city: str
    depart_time: str
    arrive_time: str
    price: float
    image_url: Optional[str] = None
    status: Optional[str] = "active"


@router.get("/buses")
def admin_get_buses(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE buses SET status = 'completed'
            WHERE status = 'active' AND depart_time < NOW()
        """))
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE b.company LIKE :s OR b.from_city LIKE :s OR b.to_city LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT b.*,
                   (SELECT img.image_url FROM images img
                    WHERE img.entity_type='bus' AND img.entity_id=b.bus_id LIMIT 1) AS image_url,
                   (SELECT COUNT(*) FROM bus_seats bs
                    WHERE bs.bus_id = b.bus_id AND bs.is_booked = 0) AS avail_total,
                   (SELECT COUNT(*) FROM bus_seats bs
                    WHERE bs.bus_id = b.bus_id) AS total_seats,
                   (SELECT COUNT(*) FROM bus_seats bs
                    WHERE bs.bus_id = b.bus_id AND bs.is_booked = 0 AND bs.seat_class = 'standard') AS avail_standard,
                   (SELECT COUNT(*) FROM bus_seats bs
                    WHERE bs.bus_id = b.bus_id AND bs.is_booked = 0 AND bs.seat_class = 'vip') AS avail_vip,
                   (SELECT COUNT(*) FROM bus_seats bs
                    WHERE bs.bus_id = b.bus_id AND bs.is_booked = 0 AND bs.seat_class = 'sleeper') AS avail_sleeper
            FROM buses b
            {where}
            ORDER BY b.bus_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/buses")
def admin_create_bus(data: BusRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO buses (company, from_city, to_city, depart_time, arrive_time, price)
            VALUES (:company, :from_city, :to_city, :depart, :arrive, :price)
        """), {
            "company": data.company, "from_city": data.from_city, "to_city": data.to_city,
            "depart": data.depart_time, "arrive": data.arrive_time,
            "price": data.price,
        })
        bus_id = result.lastrowid
        _save_image(conn, "bus", bus_id, data.image_url)
    return {"bus_id": bus_id, "message": "Tạo xe khách thành công"}


@router.put("/buses/{bus_id}")
def admin_update_bus(bus_id: int, data: BusRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE buses SET company=:company, from_city=:from_city, to_city=:to_city,
            depart_time=:depart, arrive_time=:arrive, price=:price,
            status=:status WHERE bus_id=:id
        """), {
            "company": data.company, "from_city": data.from_city, "to_city": data.to_city,
            "depart": data.depart_time, "arrive": data.arrive_time,
            "price": data.price,
            "status": data.status or "active", "id": bus_id
        })
        _save_image(conn, "bus", bus_id, data.image_url)
    return {"message": "Cập nhật xe khách thành công"}


@router.delete("/buses/{bus_id}")
def admin_delete_bus(bus_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM images WHERE entity_type='bus' AND entity_id=:id"), {"id": bus_id})
        conn.execute(text("DELETE FROM bus_seats WHERE bus_id = :id"), {"id": bus_id})
        conn.execute(text("DELETE FROM buses WHERE bus_id = :id"), {"id": bus_id})
    return {"message": "Xóa xe khách thành công"}


# ── REVIEWS ──────────────────────────────────────────────────────
@router.get("/reviews")
def admin_get_reviews(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE u.full_name LIKE :s OR u.email LIKE :s OR r.comment LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT r.*,
                   u.full_name, u.email,
                   CASE
                       WHEN r.entity_type = 'hotel' THEN (SELECT h.name FROM hotels h WHERE h.hotel_id = r.entity_id)
                       ELSE NULL
                   END AS entity_name
            FROM reviews r
            LEFT JOIN users u ON u.user_id = r.user_id
            {where}
            ORDER BY r.created_at DESC, r.review_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(row._mapping) for row in rows]


@router.delete("/reviews/{review_id}")
def admin_delete_review(review_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM reviews WHERE review_id = :id"), {"id": review_id})
    return {"message": "Xóa đánh giá thành công"}


# ── WITHDRAWAL REQUESTS ──────────────────────────────────────────
@router.get("/withdrawals")
def admin_list_withdrawals(admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT wr.*, u.full_name, u.email,
                   a.full_name AS approved_by_name
            FROM withdrawal_requests wr
            JOIN users u ON u.user_id = wr.user_id
            LEFT JOIN users a ON a.user_id = wr.approved_by
            ORDER BY wr.created_at DESC
        """)).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/withdrawals/{withdrawal_id}/approve")
async def admin_approve_withdrawal(
    withdrawal_id: int,
    background_tasks: BackgroundTasks,
    admin_id: int = Depends(get_admin_user),
):
    from email_service import send_withdrawal_success_email

    user_email = None
    user_name = None
    amount = None

    with engine.begin() as conn:
        req = conn.execute(
            text("SELECT wr.*, u.full_name, u.email FROM withdrawal_requests wr JOIN users u ON u.user_id = wr.user_id WHERE wr.wr_id = :id FOR UPDATE"),
            {"id": withdrawal_id}
        ).fetchone()
        if not req:
            raise HTTPException(404, "Không tìm thấy yêu cầu")
        req_dict = dict(req._mapping)
        if req_dict["status"] != "pending":
            raise HTTPException(400, "Yêu cầu này đã được xử lý")

        balance = conn.execute(
            text("SELECT wallet FROM users WHERE user_id = :uid FOR UPDATE"),
            {"uid": req_dict["user_id"]}
        ).scalar()
        if float(balance or 0) < float(req_dict["amount"]):
            raise HTTPException(400, "Số dư người dùng không đủ")

        # Trừ ví
        conn.execute(
            text("UPDATE users SET wallet = wallet - :amount WHERE user_id = :uid"),
            {"amount": req_dict["amount"], "uid": req_dict["user_id"]}
        )
        # Cập nhật trạng thái + lưu admin đã duyệt
        conn.execute(
            text("UPDATE withdrawal_requests SET status = 'completed', approved_by = :aid WHERE wr_id = :id"),
            {"id": withdrawal_id, "aid": admin_id}
        )
        # Ghi lịch sử
        conn.execute(text("""
            INSERT INTO wallet_transactions (user_id, amount, type, description, status)
            VALUES (:uid, :amount, 'withdrawal', :desc, 'success')
        """), {
            "uid": req_dict["user_id"],
            "amount": -float(req_dict["amount"]),
            "desc": f"Rút tiền về {req_dict['bank_name']} - TK {req_dict['account_no']}",
        })

        user_email = req_dict["email"]
        user_name = req_dict["full_name"]
        amount = float(req_dict["amount"])
        bank_info = f"{req_dict['bank_name']} - {req_dict['account_no']} ({req_dict['account_name']})"

    background_tasks.add_task(
        send_withdrawal_success_email, user_email, user_name, amount, bank_info
    )
    return {"success": True, "message": "Đã xác nhận chuyển tiền thành công"}


@router.post("/withdrawals/{withdrawal_id}/reject")
def admin_reject_withdrawal(withdrawal_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        req = conn.execute(
            text("SELECT status FROM withdrawal_requests WHERE wr_id = :id"),
            {"id": withdrawal_id}
        ).fetchone()
        if not req:
            raise HTTPException(404, "Không tìm thấy yêu cầu")
        if dict(req._mapping)["status"] != "pending":
            raise HTTPException(400, "Yêu cầu này đã được xử lý")
        conn.execute(
            text("UPDATE withdrawal_requests SET status = 'rejected', approved_by = :aid WHERE wr_id = :id"),
            {"id": withdrawal_id, "aid": admin_id}
        )
    return {"success": True, "message": "Đã từ chối yêu cầu rút tiền"}


# ── WALLET MANAGEMENT ─────────────────────────────────────────────
@router.get("/wallets")
def admin_list_wallets(admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT user_id, full_name, email, wallet, created_at
            FROM users
            ORDER BY wallet DESC
        """)).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/wallets/{user_id}/transactions")
def admin_wallet_transactions(user_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT
                wt.transaction_id,
                wt.amount,
                wt.type,
                wt.description,
                wt.status,
                wt.created_at,
                -- Với giao dịch rút tiền: lấy thông tin admin đã duyệt
                CASE wt.type
                    WHEN 'withdrawal' THEN (
                        SELECT u2.full_name
                        FROM withdrawal_requests wr2
                        JOIN users u2 ON u2.user_id = wr2.approved_by
                        WHERE wr2.user_id = wt.user_id
                          AND wr2.status IN ('completed','rejected')
                          AND ABS(wr2.amount - ABS(wt.amount)) < 1
                          AND wr2.updated_at >= DATE_SUB(wt.created_at, INTERVAL 1 MINUTE)
                          AND wr2.updated_at <= DATE_ADD(wt.created_at, INTERVAL 1 MINUTE)
                        LIMIT 1
                    )
                    ELSE NULL
                END AS approved_by_name
            FROM wallet_transactions wt
            WHERE wt.user_id = :uid
            ORDER BY wt.created_at DESC
        """), {"uid": user_id}).fetchall()
    return [dict(r._mapping) for r in rows]


# ── PROMOTIONS ───────────────────────────────────────────────────
class PromotionRequest(BaseModel):
    code: str
    description: Optional[str] = None
    discount_type: str = "percent"
    discount_percent: Optional[float] = 0
    max_discount: Optional[float] = 0
    min_order_value: Optional[float] = 0
    usage_limit: Optional[int] = 100
    per_user_limit: Optional[int] = None
    applies_to: Optional[str] = "all"
    status: Optional[str] = "active"
    expired_at: Optional[str] = None


@router.get("/promotions")
def admin_get_promotions(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        if search:
            params["s"] = f"%{search}%"
            rows = conn.execute(text(
                "SELECT * FROM promotions WHERE code LIKE :s OR description LIKE :s ORDER BY promo_id DESC LIMIT :limit OFFSET :skip"
            ), params).fetchall()
        else:
            rows = conn.execute(text(
                "SELECT * FROM promotions ORDER BY promo_id DESC LIMIT :limit OFFSET :skip"
            ), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/promotions")
def admin_create_promotion(data: PromotionRequest, admin_id: int = Depends(get_admin_user)):
    code = data.code.strip().upper()
    with engine.begin() as conn:
        exists = conn.execute(
            text("SELECT promo_id FROM promotions WHERE code = :code"), {"code": code}
        ).fetchone()
        if exists:
            raise HTTPException(400, f"Mã '{code}' đã tồn tại")
        result = conn.execute(text("""
            INSERT INTO promotions
                (code, description, discount_type, discount_percent, max_discount,
                 min_order_value, usage_limit, per_user_limit, applies_to, status, expired_at)
            VALUES
                (:code, :desc, :dtype, :dpct, :maxd, :minv, :ulimit, :per_user, :applies, :status, :exp)
        """), {
            "code": code, "desc": data.description, "dtype": data.discount_type,
            "dpct": data.discount_percent or 0, "maxd": data.max_discount or 0,
            "minv": data.min_order_value or 0, "ulimit": data.usage_limit or 100,
            "per_user": data.per_user_limit,
            "applies": data.applies_to or "all", "status": data.status or "active",
            "exp": data.expired_at,
        })
        promo_id = result.lastrowid

        # Gửi thông báo đến tất cả user
        discount_str = (
            f"{data.discount_percent}%" if data.discount_type == "percent"
            else f"{int(data.max_discount or 0):,}₫"
        )
        title = f"🎁 Mã giảm giá mới: {code}"
        content = f"Giảm {discount_str}{' - ' + data.description if data.description else ''}. Hạn dùng đến {str(data.expired_at)[:10]}."
        user_ids = conn.execute(
            text("SELECT user_id FROM users")
        ).fetchall()
        for row in user_ids:
            create_notification(conn, row.user_id, type="promotion", title=title, content=content, related_id=promo_id)

    return {"promo_id": promo_id, "message": "Tạo mã giảm giá thành công"}


@router.put("/promotions/{promo_id}")
def admin_update_promotion(promo_id: int, data: PromotionRequest, admin_id: int = Depends(get_admin_user)):
    code = data.code.strip().upper()
    with engine.begin() as conn:
        # Check duplicate code (excluding self)
        exists = conn.execute(
            text("SELECT promo_id FROM promotions WHERE code = :code AND promo_id != :id"),
            {"code": code, "id": promo_id}
        ).fetchone()
        if exists:
            raise HTTPException(400, f"Mã '{code}' đã được dùng bởi promotion khác")
        conn.execute(text("""
            UPDATE promotions SET
                code=:code, description=:desc, discount_type=:dtype,
                discount_percent=:dpct, max_discount=:maxd,
                min_order_value=:minv, usage_limit=:ulimit, per_user_limit=:per_user,
                applies_to=:applies, status=:status, expired_at=:exp
            WHERE promo_id=:id
        """), {
            "code": code, "desc": data.description, "dtype": data.discount_type,
            "dpct": data.discount_percent or 0, "maxd": data.max_discount or 0,
            "minv": data.min_order_value or 0, "ulimit": data.usage_limit or 100,
            "per_user": data.per_user_limit,
            "applies": data.applies_to or "all", "status": data.status or "active",
            "exp": data.expired_at, "id": promo_id,
        })
    return {"message": "Cập nhật mã giảm giá thành công"}


@router.delete("/promotions/{promo_id}")
def admin_delete_promotion(promo_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM promotions WHERE promo_id = :id"), {"id": promo_id})
    return {"message": "Xóa mã giảm giá thành công"}


@router.put("/promotions/{promo_id}/toggle")
def admin_toggle_promotion(promo_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        promo = conn.execute(
            text("SELECT status FROM promotions WHERE promo_id = :id"), {"id": promo_id}
        ).fetchone()
        if not promo:
            raise HTTPException(404, "Không tìm thấy mã giảm giá")
        new_status = "inactive" if dict(promo._mapping)["status"] == "active" else "active"
        conn.execute(
            text("UPDATE promotions SET status=:s WHERE promo_id=:id"),
            {"s": new_status, "id": promo_id}
        )
    return {"status": new_status}


# ── BANNERS ──────────────────────────────────────────────────────
class BannerRequest(BaseModel):
    title: Optional[str] = None
    subtitle: Optional[str] = None
    image_url: str
    link_url: Optional[str] = None
    display_order: Optional[int] = 0
    is_active: Optional[int] = 1
    page_display: Optional[str] = "home"
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@router.get("/banners")
def admin_get_banners(admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT * FROM banners ORDER BY display_order ASC, banner_id DESC"
        )).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/banners")
def admin_create_banner(data: BannerRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text("""
            INSERT INTO banners (title, subtitle, image_url, link_url, display_order, is_active, page_display, start_date, end_date)
            VALUES (:title, :subtitle, :image_url, :link_url, :order, :active, :page_display, :start, :end)
        """), {
            "title": data.title, "subtitle": data.subtitle,
            "image_url": data.image_url, "link_url": data.link_url,
            "order": data.display_order or 0, "active": data.is_active if data.is_active is not None else 1,
            "page_display": data.page_display or "home",
            "start": data.start_date or None, "end": data.end_date or None,
        })
    return {"banner_id": result.lastrowid, "message": "Tạo banner thành công"}


@router.put("/banners/{banner_id}")
def admin_update_banner(banner_id: int, data: BannerRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE banners SET
                title=:title, subtitle=:subtitle, image_url=:image_url,
                link_url=:link_url, display_order=:order,
                is_active=:active, page_display=:page_display, start_date=:start, end_date=:end
            WHERE banner_id=:id
        """), {
            "title": data.title, "subtitle": data.subtitle,
            "image_url": data.image_url, "link_url": data.link_url,
            "order": data.display_order or 0, "active": data.is_active if data.is_active is not None else 1,
            "page_display": data.page_display or "home",
            "start": data.start_date or None, "end": data.end_date or None, "id": banner_id,
        })
    return {"message": "Cập nhật banner thành công"}


@router.delete("/banners/{banner_id}")
def admin_delete_banner(banner_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM banners WHERE banner_id=:id"), {"id": banner_id})
    return {"message": "Xóa banner thành công"}


@router.put("/banners/{banner_id}/toggle")
def admin_toggle_banner(banner_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        current = conn.execute(
            text("SELECT is_active FROM banners WHERE banner_id=:id"), {"id": banner_id}
        ).fetchone()
        if not current:
            raise HTTPException(404, "Không tìm thấy banner")
        new_val = 0 if dict(current._mapping)["is_active"] else 1
        conn.execute(
            text("UPDATE banners SET is_active=:v WHERE banner_id=:id"),
            {"v": new_val, "id": banner_id}
        )
    return {"is_active": new_val}


# ── DESTINATIONS MANAGEMENT ─────────────────────────────────────
class DestinationRequest(BaseModel):
    city: str
    name: str
    description: Optional[str] = None
    image_url: Optional[str] = None


@router.get("/destinations")
def admin_get_destinations(admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        # Migration: xóa các cột cũ không còn dùng nữa
        for col in ["image_url", "avg_rating", "review_count", "img_url"]:
            try:
                conn.execute(text(f"ALTER TABLE destinations DROP COLUMN {col}"))
                conn.commit()
            except Exception:
                pass
        rows = conn.execute(text("""
            SELECT d.*,
                   (SELECT img.image_url FROM images img
                    WHERE img.entity_type = 'destination' AND img.entity_id = d.destination_id
                    LIMIT 1) AS image_url
            FROM destinations d
            ORDER BY d.city ASC
        """)).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/destinations")
def admin_create_destination(data: DestinationRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        result = conn.execute(text(
            "INSERT INTO destinations (city, name, description) VALUES (:city, :name, :desc)"
        ), {"city": data.city, "name": data.name, "desc": data.description})
        dest_id = result.lastrowid
        _save_image(conn, "destination", dest_id, data.image_url)
    return {"destination_id": dest_id, "message": "Tạo địa điểm thành công"}


@router.put("/destinations/{dest_id}")
def admin_update_destination(dest_id: int, data: DestinationRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text(
            "UPDATE destinations SET city=:city, name=:name, description=:desc WHERE destination_id=:id"
        ), {"city": data.city, "name": data.name, "desc": data.description, "id": dest_id})
        _save_image(conn, "destination", dest_id, data.image_url)
    return {"message": "Cập nhật địa điểm thành công"}


@router.delete("/destinations/{dest_id}")
def admin_delete_destination(dest_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM images WHERE entity_type='destination' AND entity_id=:id"), {"id": dest_id})
        conn.execute(text("DELETE FROM destinations WHERE destination_id=:id"), {"id": dest_id})
    return {"message": "Xóa địa điểm thành công"}


# ── BOOKING MODIFICATIONS MANAGEMENT ───────────────────────────
@router.get("/modifications")
def get_modifications(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT m.*,
                u.full_name AS user_name, u.email AS user_email,
                CASE bi.entity_type
                    WHEN 'room'   THEN CONCAT(h.name, ' - ', rt.name)
                    WHEN 'flight' THEN CONCAT(f.airline, ': ', f.from_city, ' → ', f.to_city)
                    WHEN 'bus'    THEN CONCAT(bs.company, ': ', bs.from_city, ' → ', bs.to_city)
                    ELSE CONCAT(bi.entity_type, ' #', bi.entity_id)
                END AS entity_name,
                bi.entity_type,
                adm.full_name AS approved_by_name
            FROM booking_modifications m
            JOIN users u ON u.user_id = m.user_id
            JOIN booking_items bi ON bi.booking_id = m.booking_id
            LEFT JOIN room_types rt ON rt.room_type_id = bi.entity_id AND bi.entity_type = 'room'
            LEFT JOIN hotels h ON h.hotel_id = rt.hotel_id
            LEFT JOIN flights f ON f.flight_id = bi.entity_id AND bi.entity_type = 'flight'
            LEFT JOIN buses bs ON bs.bus_id = bi.entity_id AND bi.entity_type = 'bus'
            LEFT JOIN users adm ON adm.user_id = m.approved_by
            ORDER BY m.created_at DESC
            LIMIT :limit OFFSET :skip
        """), {"limit": limit, "skip": skip}).fetchall()
        return [dict(r._mapping) for r in rows]


@router.post("/modifications/{mod_id}/approve")
def approve_modification(mod_id: int, data: dict = {}, background_tasks: BackgroundTasks = None,
                         admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        mod = conn.execute(
            text("SELECT * FROM booking_modifications WHERE mod_id = :id AND status = 'pending' FOR UPDATE"),
            {"id": mod_id}
        ).fetchone()
        if not mod:
            raise HTTPException(404, "Yêu cầu không tồn tại")
        m = dict(mod._mapping)

        refund_amount = float(m["refund_amount"] or 0)
        refund_method = m["refund_method"] or "wallet"
        booking_id    = m["booking_id"]
        user_id       = m["user_id"]

        if refund_amount > 0 and refund_method == "wallet":
            conn.execute(
                text("UPDATE users SET wallet = wallet + :amt WHERE user_id = :uid"),
                {"amt": refund_amount, "uid": user_id}
            )
            conn.execute(text("""
                INSERT INTO wallet_transactions (user_id, amount, type, description, status)
                VALUES (:uid, :amt, 'refund', :desc, 'success')
            """), {"uid": user_id, "amt": refund_amount,
                   "desc": f"Hoàn tiền {'hủy' if m['type'] == 'cancel' else 'đổi lịch'} #{booking_id}"})

        # Nếu là đổi lịch: cập nhật booking, booking_items và transfer ghế
        if m["type"] == "reschedule" and float(m.get("new_price") or 0) > 0:
            new_entity_id  = m.get("new_entity_id")
            new_seat_class = m.get("new_seat_class")
            new_price      = float(m["new_price"])

            # Lấy entity_type và thông tin ghế cũ để transfer
            item_row = conn.execute(
                text("SELECT entity_type, entity_id, seat_class, quantity FROM booking_items WHERE booking_id = :bid LIMIT 1"),
                {"bid": booking_id}
            ).fetchone()
            if item_row:
                item_d     = dict(item_row._mapping)
                entity_type    = item_d["entity_type"]
                old_entity_id  = item_d["entity_id"]
                old_seat_class = item_d.get("seat_class") or new_seat_class or ""
                quantity       = item_d.get("quantity") or 1

                # Transfer ghế nếu là vận tải và có thay đổi
                if entity_type in ("flight", "bus", "train") and new_entity_id and new_seat_class:
                    from routers.bookings import _transfer_seats
                    try:
                        _transfer_seats(conn, entity_type, old_entity_id, new_entity_id,
                                        old_seat_class, new_seat_class, quantity)
                    except Exception:
                        pass  # Ghế đã được chuyển khi user thanh toán

            if new_entity_id:
                conn.execute(
                    text("UPDATE booking_items SET entity_id = :eid WHERE booking_id = :bid"),
                    {"eid": new_entity_id, "bid": booking_id},
                )
            if new_seat_class:
                conn.execute(
                    text("UPDATE booking_items SET seat_class = :sc WHERE booking_id = :bid"),
                    {"sc": new_seat_class, "bid": booking_id},
                )
            if m.get("new_check_in"):
                conn.execute(
                    text("UPDATE booking_items SET check_in_date = :ci, check_out_date = :co WHERE booking_id = :bid"),
                    {"ci": str(m["new_check_in"]), "co": str(m["new_check_out"]), "bid": booking_id},
                )
            conn.execute(
                text("UPDATE bookings SET total_price = :np, final_amount = :np WHERE booking_id = :bid"),
                {"np": new_price, "bid": booking_id},
            )
            conn.execute(
                text("UPDATE booking_items SET price = :np WHERE booking_id = :bid"),
                {"np": new_price, "bid": booking_id},
            )

        conn.execute(
            text("UPDATE booking_modifications SET status = 'approved', admin_note = :note, approved_by = :ab, approved_at = NOW() WHERE mod_id = :id"),
            {"id": mod_id, "note": (data or {}).get("note", ""), "ab": admin_id}
        )

    return {"message": "Đã duyệt yêu cầu"}


@router.post("/modifications/{mod_id}/reject")
def reject_modification(mod_id: int, data: dict = {}, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        mod = conn.execute(
            text("SELECT * FROM booking_modifications WHERE mod_id = :id AND status = 'pending'"),
            {"id": mod_id}
        ).fetchone()
        if not mod:
            raise HTTPException(404, "Yêu cầu không tồn tại")
        m = dict(mod._mapping)

        # Nếu là reschedule pending (cần thanh toán thêm nhưng chưa trả), huỷ mod thôi
        # Không restore cancel vì booking đã bị huỷ ngay khi user gửi yêu cầu

        conn.execute(
            text("UPDATE booking_modifications SET status = 'rejected', admin_note = :note, approved_by = :ab, approved_at = NOW() WHERE mod_id = :id"),
            {"id": mod_id, "note": (data or {}).get("note", ""), "ab": admin_id}
        )

    return {"message": "Đã từ chối yêu cầu"}


# ── TRAINS ───────────────────────────────────────────────────────
class TrainRequest(BaseModel):
    train_code: str
    from_city: str
    to_city: str
    from_station: str
    to_station: str
    depart_time: str
    arrive_time: str
    status: Optional[str] = "active"


@router.get("/trains")
def admin_get_trains(skip: int = 0, limit: int = 50, search: str = "", admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE trains SET status = 'completed'
            WHERE status = 'active' AND depart_time < NOW()
        """))
    with engine.connect() as conn:
        params: dict = {"limit": limit, "skip": skip}
        where = ""
        if search:
            where = "WHERE t.train_code LIKE :s OR t.from_city LIKE :s OR t.to_city LIKE :s"
            params["s"] = f"%{search}%"
        rows = conn.execute(text(f"""
            SELECT t.*,
                   TIMESTAMPDIFF(MINUTE, t.depart_time, t.arrive_time) AS duration_minutes,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id AND ts.is_booked = 0) AS avail_total,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id) AS total_seats,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id AND ts.is_booked = 0 AND ts.seat_class = 'hard_seat') AS avail_hard_seat,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id AND ts.is_booked = 0 AND ts.seat_class = 'soft_seat') AS avail_soft_seat,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id AND ts.is_booked = 0 AND ts.seat_class = 'hard_sleeper') AS avail_hard_sleeper,
                   (SELECT COUNT(*) FROM train_seats ts WHERE ts.train_id = t.train_id AND ts.is_booked = 0 AND ts.seat_class = 'soft_sleeper') AS avail_soft_sleeper
            FROM trains t
            {where}
            ORDER BY t.depart_time DESC, t.train_id DESC
            LIMIT :limit OFFSET :skip
        """), params).fetchall()
    return [dict(r._mapping) for r in rows]


@router.post("/trains")
def admin_create_train(data: TrainRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        existing = conn.execute(
            text("SELECT train_id FROM trains WHERE train_code = :code AND DATE(depart_time) = DATE(:depart)"),
            {"code": data.train_code, "depart": data.depart_time}
        ).fetchone()
        if existing:
            raise HTTPException(400, "Tàu này đã tồn tại trong ngày đó")
        result = conn.execute(text("""
            INSERT INTO trains (train_code, from_city, to_city, from_station, to_station,
                                depart_time, arrive_time, status)
            VALUES (:code, :from_city, :to_city, :from_st, :to_st, :depart, :arrive, :status)
        """), {
            "code": data.train_code, "from_city": data.from_city, "to_city": data.to_city,
            "from_st": data.from_station, "to_st": data.to_station,
            "depart": data.depart_time, "arrive": data.arrive_time, "status": data.status,
        })
    return {"train_id": result.lastrowid, "message": "Tạo chuyến tàu thành công"}


@router.put("/trains/{train_id}")
def admin_update_train(train_id: int, data: TrainRequest, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE trains SET train_code=:code, from_city=:from_city, to_city=:to_city,
            from_station=:from_st, to_station=:to_st,
            depart_time=:depart, arrive_time=:arrive, status=:status
            WHERE train_id=:id
        """), {
            "code": data.train_code, "from_city": data.from_city, "to_city": data.to_city,
            "from_st": data.from_station, "to_st": data.to_station,
            "depart": data.depart_time, "arrive": data.arrive_time,
            "status": data.status, "id": train_id,
        })
    return {"message": "Cập nhật chuyến tàu thành công"}


@router.delete("/trains/{train_id}")
def admin_delete_train(train_id: int, admin_id: int = Depends(get_admin_user)):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM train_seats WHERE train_id = :id"), {"id": train_id})
        conn.execute(text("DELETE FROM trains WHERE train_id = :id"), {"id": train_id})
    return {"message": "Xóa chuyến tàu thành công"}


# ── DOWNLOAD EXCEL TEMPLATE ─────────────────────────────────────────
from fastapi.responses import StreamingResponse as _StreamingResponse

EXCEL_TEMPLATES = {
    "hotels": {
        # Bắt buộc: name | Tùy chọn: address, city, description, amenities
        "headers": ["name *", "address", "city", "description", "amenities"],
        "sample": [
            ["Khách sạn Mặt Trời", "12 Trần Phú, Đà Nẵng", "Đà Nẵng", "Khách sạn 4 sao view biển", "Wifi, Hồ bơi, Gym"],
            ["Villa Hội An", "45 Nguyễn Thị Minh Khai, Hội An", "Hội An", "Villa cổ kính yên tĩnh", "Wifi, Bể bơi riêng"],
        ],
    },
    "flights": {
        # Bắt buộc: airline, from_city, to_city, price | Tùy chọn: depart_time, arrive_time
        "headers": ["airline *", "from_city *", "to_city *", "price *", "depart_time", "arrive_time"],
        "sample": [
            ["Vietnam Airlines", "Hà Nội", "Đà Nẵng", 850000, "2026-05-01 07:00", "2026-05-01 08:20", 180],
            ["VietJet Air", "Hồ Chí Minh", "Phú Quốc", 650000, "2026-05-10 09:00", "2026-05-10 10:00", 200],
        ],
    },
    "buses": {
        # Bắt buộc: company, from_city, to_city, price | Tùy chọn: depart_time, arrive_time
        "headers": ["company *", "from_city *", "to_city *", "price *", "depart_time", "arrive_time"],
        "sample": [
            ["Phương Trang", "Hà Nội", "Đà Lạt", 350000, "2026-05-01 19:00", "2026-05-02 09:00", 40],
            ["Kumho Samco", "Hồ Chí Minh", "Vũng Tàu", 120000, "2026-05-05 08:00", "2026-05-05 10:30", 35],
        ],
    },
    "trains": {
        # Bắt buộc: train_code, from_city, to_city, from_station, to_station, depart_time, arrive_time, price
        "headers": ["train_code *", "from_city *", "to_city *", "from_station *", "to_station *", "depart_time *", "arrive_time *", "price *", "status"],
        "sample": [
            ["SE1", "Hà Nội", "Hồ Chí Minh", "Ga Hà Nội", "Ga Sài Gòn", "2026-05-01 06:00", "2026-05-02 12:00", 700000, "active"],
            ["SE3", "Hà Nội", "Đà Nẵng", "Ga Hà Nội", "Ga Đà Nẵng", "2026-05-02 07:30", "2026-05-02 20:30", 420000, "active"],
        ],
    },
    "promotions": {
        # Bắt buộc: code, discount_percent | Tùy chọn: description, discount_type, max_discount, min_order_value, usage_limit, applies_to
        "headers": ["code *", "discount_percent *", "description", "discount_type", "max_discount", "min_order_value", "usage_limit", "applies_to"],
        "sample": [
            ["SUMMER20", 20, "Giảm 20% mùa hè", "percent", 500000, 1000000, 100, "all"],
            ["HOTEL50K", 50000, "Giảm 50k khách sạn", "fixed", "", 500000, 50, "hotel"],
        ],
    },
}

@router.get("/template/{section}")
def download_template(section: str, admin_id: int = Depends(get_admin_user)):
    """Tải file Excel mẫu cho từng section."""
    if section not in EXCEL_TEMPLATES:
        raise HTTPException(400, f"Không có template cho '{section}'")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "Thiếu thư viện openpyxl. Chạy: pip install openpyxl")

    tmpl = EXCEL_TEMPLATES[section]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = section.capitalize()

    # Header row — màu xanh đậm, chữ trắng (ARGB format: FF + hex)
    header_fill = PatternFill("solid", fgColor="FF0052CC")
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    for col, h in enumerate(tmpl["headers"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(18, len(h) + 4)

    # Sample rows — màu xanh nhạt xen kẽ
    alt_fill = PatternFill("solid", fgColor="FFE8F0FE")
    for row_idx, sample_row in enumerate(tmpl["sample"], 2):
        for col_idx, val in enumerate(sample_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return _StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_{section}.xlsx"},
    )


# ── IMPORT FROM EXCEL ───────────────────────────────────────────────
@router.post("/import/{section}")
async def import_excel(
    section: str,
    file: UploadFile = File(...),
    admin_id: int = Depends(get_admin_user),
):
    ALLOWED = {"hotels", "flights", "buses", "trains", "promotions"}
    if section not in ALLOWED:
        raise HTTPException(400, f"Không hỗ trợ import cho section '{section}'")
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "Thiếu thư viện openpyxl. Chạy: pip install openpyxl")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "File Excel không hợp lệ hoặc bị lỗi")

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    inserted = 0
    skipped = 0
    errors = []

    def get(row_dict, *keys, default=None):
        for k in keys:
            v = row_dict.get(k)
            if v is not None and str(v).strip() != "":
                return v
        return default

    with engine.begin() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            if all(v is None or str(v).strip() == "" for v in row_dict.values()):
                continue
            try:
                # Strip " *" khỏi key (do header có dấu * để đánh dấu bắt buộc)
                row_dict = {k.replace(" *", "").strip(): v for k, v in row_dict.items()}

                if section == "hotels":
                    name = get(row_dict, "name")
                    if not name: skipped += 1; continue
                    if conn.execute(text("SELECT hotel_id FROM hotels WHERE name=:n LIMIT 1"), {"n": str(name)}).fetchone():
                        skipped += 1; continue
                    dest_id = None
                    city_name = get(row_dict, "city")
                    if city_name:
                        r = conn.execute(text("SELECT destination_id FROM destinations WHERE city LIKE :c LIMIT 1"), {"c": f"%{city_name}%"}).fetchone()
                        if r: dest_id = r[0]
                    # amenities là JSON array
                    amen_raw = get(row_dict, "amenities") or ""
                    import json as _json
                    amen_json = _json.dumps([a.strip() for a in str(amen_raw).split(",") if a.strip()]) if amen_raw else None
                    conn.execute(text("""
                        INSERT INTO hotels (name, address, destination_id, description, amenities)
                        VALUES (:name, :addr, :dest, :desc, :amen)
                    """), {
                        "name": str(name),
                        "addr": str(get(row_dict, "address") or ""),
                        "dest": int(dest_id) if dest_id else None,
                        "desc": str(get(row_dict, "description") or ""),
                        "amen": amen_json,
                    })
                    inserted += 1

                elif section == "flights":
                    airline   = get(row_dict, "airline")
                    from_city = get(row_dict, "from_city")
                    to_city   = get(row_dict, "to_city")
                    price     = get(row_dict, "price")
                    if not all([airline, from_city, to_city, price]): skipped += 1; continue
                    conn.execute(text("""
                        INSERT INTO flights (airline, from_city, to_city, depart_time, arrive_time, price)
                        VALUES (:airline, :from_c, :to_c, :dep, :arr, :price)
                    """), {
                        "airline": str(airline), "from_c": str(from_city), "to_c": str(to_city),
                        "dep":   get(row_dict, "depart_time") or None,
                        "arr":   get(row_dict, "arrive_time") or None,
                        "price": float(price),
                    })
                    inserted += 1

                elif section == "buses":
                    company   = get(row_dict, "company")
                    from_city = get(row_dict, "from_city")
                    to_city   = get(row_dict, "to_city")
                    price     = get(row_dict, "price")
                    if not all([company, from_city, to_city, price]): skipped += 1; continue
                    conn.execute(text("""
                        INSERT INTO buses (company, from_city, to_city, depart_time, arrive_time, price)
                        VALUES (:company, :from_c, :to_c, :dep, :arr, :price)
                    """), {
                        "company": str(company), "from_c": str(from_city), "to_c": str(to_city),
                        "dep":   get(row_dict, "depart_time") or None,
                        "arr":   get(row_dict, "arrive_time") or None,
                        "price": float(price),
                    })
                    inserted += 1

                elif section == "trains":
                    code      = get(row_dict, "train_code")
                    from_city = get(row_dict, "from_city")
                    to_city   = get(row_dict, "to_city")
                    from_st   = get(row_dict, "from_station")
                    to_st     = get(row_dict, "to_station")
                    dep       = get(row_dict, "depart_time")
                    arr       = get(row_dict, "arrive_time")
                    price     = get(row_dict, "price")
                    if not all([code, from_city, to_city, from_st, to_st, dep, arr, price]):
                        skipped += 1; continue
                    conn.execute(text("""
                        INSERT INTO trains (train_code, from_city, to_city, from_station, to_station, depart_time, arrive_time, price, status)
                        VALUES (:code, :from_c, :to_c, :from_st, :to_st, :dep, :arr, :price, :status)
                    """), {
                        "code": str(code), "from_c": str(from_city), "to_c": str(to_city),
                        "from_st": str(from_st), "to_st": str(to_st),
                        "dep": str(dep), "arr": str(arr),
                        "price": float(price),
                        "status": str(get(row_dict, "status") or "active"),
                    })
                    inserted += 1

                elif section == "promotions":
                    code     = get(row_dict, "code")
                    discount = get(row_dict, "discount_percent")
                    if not code or not discount: skipped += 1; continue
                    if conn.execute(text("SELECT promo_id FROM promotions WHERE code=:c LIMIT 1"), {"c": str(code)}).fetchone():
                        skipped += 1; continue
                    conn.execute(text("""
                        INSERT INTO promotions (code, description, discount_type, discount_percent,
                                               max_discount, min_order_value, usage_limit, applies_to, status)
                        VALUES (:code, :desc, :dtype, :dpct, :maxd, :minv, :ulim, :apto, :status)
                    """), {
                        "code":  str(code),
                        "desc":  str(get(row_dict, "description", "Mô tả") or ""),
                        "dtype": str(get(row_dict, "discount_type") or "percent"),
                        "dpct":  float(discount),
                        "maxd":  float(get(row_dict, "max_discount") or 0) or None,
                        "minv":  float(get(row_dict, "min_order_value") or 0) or None,
                        "ulim":  int(get(row_dict, "usage_limit") or 100),
                        "apto":  str(get(row_dict, "applies_to") or "all"),
                        "status": str(get(row_dict, "status") or "active"),
                    })
                    inserted += 1

            except Exception as e:
                errors.append(f"Hàng {row_idx}: {str(e)[:120]}")
                if len(errors) >= 20:
                    break

    return {"inserted": inserted, "skipped": skipped, "errors": errors}

